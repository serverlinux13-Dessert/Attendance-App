
import base64
import hashlib
import hmac
import ipaddress
import io
import json
import math
import os
import secrets
import sqlite3
import uuid
from datetime import date, datetime, time, timedelta, timezone
from functools import wraps
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import qrcode
import requests
from flask import Flask, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from flask_cors import CORS
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "database.db"
PUBLIC_DIR = BASE_DIR / "public"

try:
    IST = ZoneInfo("Asia/Kolkata")
except ZoneInfoNotFoundError:
    IST = timezone(timedelta(hours=5, minutes=30))

QR_TTL_SECONDS = 300
OTP_LENGTH = 6
OTP_MAX_ATTEMPTS = 5
BREAK_CUTOFF_HOUR = 4
DEFAULT_EMPLOYEE_PIN = "1111"
DEFAULT_ADMIN_PIN = "1234"
APP_ENV = os.getenv("APP_ENV", os.getenv("FLASK_ENV", "development")).strip().lower()
IS_PRODUCTION = APP_ENV == "production"
SECRET_KEY = os.getenv("SECRET_KEY", "").strip()
if not SECRET_KEY:
    if IS_PRODUCTION:
        raise RuntimeError("SECRET_KEY environment variable is required when APP_ENV=production")
    SECRET_KEY = "dev-change-this-secret"
REQUIRE_OFFICE_NETWORK = os.getenv("REQUIRE_OFFICE_NETWORK", "0") == "1"
ALLOWED_SUBNET = os.getenv("ALLOWED_SUBNET", "").strip()
ALLOWED_SUBNETS = os.getenv("ALLOWED_SUBNETS", "").strip()
ALLOW_ADMIN_FROM_ANYWHERE = os.getenv("ALLOW_ADMIN_FROM_ANYWHERE", "1") == "1"
TRUST_PROXY = os.getenv("TRUST_PROXY", "0") == "1"
HAS_EXPLICIT_ALLOWED_NETWORKS = bool(ALLOWED_SUBNET or ALLOWED_SUBNETS)


def env_flag(name, default=False):
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def load_allowed_networks():
    configured = []
    if ALLOWED_SUBNET:
        configured.append(ALLOWED_SUBNET)
    if ALLOWED_SUBNETS:
        configured.extend(p.strip() for p in ALLOWED_SUBNETS.split(","))

    networks = []
    seen = set()
    for value in configured:
        if not value:
            continue
        try:
            net = ipaddress.ip_network(value, strict=False)
        except ValueError:
            continue
        key = net.with_prefixlen
        if key not in seen:
            seen.add(key)
            networks.append(net)
    return networks


ALLOWED_NETWORKS = load_allowed_networks()
SESSION_COOKIE_SECURE = env_flag("SESSION_COOKIE_SECURE", IS_PRODUCTION)

app = Flask(__name__)
app.config.update(
    SECRET_KEY=SECRET_KEY,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=SESSION_COOKIE_SECURE,
)
if TRUST_PROXY:
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)
CORS(app, supports_credentials=True)


def conn_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def restore_db_from_github():
    try:
        if DB_PATH.exists():
            return

        repo = os.getenv("GITHUB_REPO", "").strip()
        if not repo:
            return

        raw_url = f"https://raw.githubusercontent.com/{repo}/main/db_backup/database.db"
        response = requests.get(raw_url, timeout=20)
        if response.status_code == 200 and response.content:
            DB_PATH.write_bytes(response.content)
    except Exception:
        pass


def backup_db_to_github():
    if not DB_PATH.exists():
        return False, "Local database file not found"

    token = os.getenv("GITHUB_TOKEN", "").strip()
    repo = os.getenv("GITHUB_REPO", "").strip()
    if not token or not repo:
        return False, "GITHUB_TOKEN or GITHUB_REPO is missing"

    try:
        encoded_db = base64.b64encode(DB_PATH.read_bytes()).decode("utf-8")
    except Exception as exc:
        app.logger.exception("Failed to read database for backup")
        return False, f"Failed to read local database: {exc}"

    contents_url = f"https://api.github.com/repos/{repo}/contents/db_backup/database.db"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github+json",
    }

    sha = None
    try:
        metadata_response = requests.get(contents_url, headers=headers, timeout=20)
    except Exception as exc:
        app.logger.exception("Failed to fetch backup metadata from GitHub")
        return False, f"Failed to contact GitHub API while reading metadata: {exc}"

    if metadata_response.status_code == 200:
        try:
            sha = (metadata_response.json() or {}).get("sha")
        except Exception as exc:
            app.logger.exception("Failed to parse GitHub metadata response")
            return False, f"Invalid metadata response from GitHub: {exc}"
    elif metadata_response.status_code != 404:
        return False, f"GitHub metadata request failed with status {metadata_response.status_code}: {metadata_response.text[:200]}"

    payload = {
        "message": "Auto backup database",
        "content": encoded_db,
    }
    if sha:
        payload["sha"] = sha

    try:
        upload_response = requests.put(contents_url, headers=headers, json=payload, timeout=30)
    except Exception as exc:
        app.logger.exception("Failed to upload backup to GitHub")
        return False, f"Failed to contact GitHub API while uploading backup: {exc}"

    if upload_response.status_code not in {200, 201}:
        return False, f"GitHub backup upload failed with status {upload_response.status_code}: {upload_response.text[:200]}"

    return True, "Database backed up"


def has_col(conn, table, col):
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r["name"] == col for r in rows)


def col_default(conn, table, col):
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    for r in rows:
        if r["name"] == col:
            v = r["dflt_value"]
            if v is None:
                return None
            return str(v).strip().strip("'\"")
    return None


def ensure_col(conn, table, col, ddl):
    if not has_col(conn, table, col):
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {ddl}")


def table_exists(conn, table):
    r = conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
    return r is not None


def now_ist():
    return datetime.now(IST)


def now_iso():
    return now_ist().isoformat()


def break_effective_date_ist(dt=None):
    local_now = dt or now_ist()
    if local_now.hour < BREAK_CUTOFF_HOUR:
        return local_now.date() - timedelta(days=1)
    return local_now.date()


def epoch_ms():
    return int(now_ist().timestamp() * 1000)


def generate_otp_code():
    return f"{secrets.randbelow(10 ** OTP_LENGTH):0{OTP_LENGTH}d}"


def otp_digest(otp_code, session_id):
    payload = f"{session_id}:{otp_code}".encode("utf-8")
    return hmac.new(SECRET_KEY.encode("utf-8"), payload, hashlib.sha256).hexdigest()


def parse_dt(v):
    d = datetime.fromisoformat(v)
    return d if d.tzinfo else d.replace(tzinfo=IST)


def get_client_ip():
    f = request.headers.get("X-Forwarded-For", "")
    return f.split(",")[0].strip() if f else (request.remote_addr or "")


def office_ok():
    if not REQUIRE_OFFICE_NETWORK:
        return True
    ip = get_client_ip()
    if ip.startswith("127.") or ip == "::1":
        return True
    if HAS_EXPLICIT_ALLOWED_NETWORKS:
        try:
            client_ip = ipaddress.ip_address(ip)
        except ValueError:
            return False
        if not ALLOWED_NETWORKS:
            return False
        return any(client_ip in net for net in ALLOWED_NETWORKS)
    return ip.startswith("10.") or ip.startswith("192.168.") or ip.startswith("172.")


def office_check():
    if office_ok():
        return None
    return jsonify({"message": "Access allowed only from office network"}), 403


def role_bypasses_office_check(role):
    return ALLOW_ADMIN_FROM_ANYWHERE and role == "ADMIN"


def valid_pin(pin):
    return isinstance(pin, str) and pin.isdigit() and len(pin) == 4


def month_range():
    t = now_ist().date()
    return t.replace(day=1), t


def resolve_date_range(from_str, to_str):
    df, dt = month_range()
    if from_str:
        df = date.fromisoformat(from_str)
    if to_str:
        dt = date.fromisoformat(to_str)
    if df > dt:
        raise ValueError("from date cannot be after to date")
    return df, dt


def to_items(rows):
    return [dict(r) for r in rows]


def user_profile(conn, user_id):
    return conn.execute(
        """
        SELECT u.*, COALESCE(c.required_hours,9) required_hours,
               COALESCE(s.start_time,'09:00') shift_start, COALESCE(s.end_time,'18:00') shift_end,
               COALESCE(s.grace_minutes,15) grace_minutes,
               COALESCE(c.name,'General') category_name, COALESCE(s.name,'General Shift') shift_name
        FROM users u
        LEFT JOIN employee_categories c ON c.id=u.category_id
        LEFT JOIN shifts s ON s.id=u.shift_id
        WHERE u.id=? AND u.active=1
        """,
        (user_id,),
    ).fetchone()


def round_overtime_hours(value):
    value = max(0.0, float(value or 0.0))
    return math.floor(value * 2) / 2


def calc_metrics(login_iso, logout_iso, profile):
    login = parse_dt(login_iso).astimezone(IST)
    logout = parse_dt(logout_iso).astimezone(IST)
    total_hours = max(0, (logout - login).total_seconds() / 3600)
    st = datetime.combine(login.date(), time.fromisoformat(profile["shift_start"]), tzinfo=IST)
    assigned_shift_hours = max(0, float(profile["required_hours"] or 0))
    ot = max(0, total_hours - assigned_shift_hours)
    late = int(login > st + timedelta(minutes=int(profile["grace_minutes"] or 0)))
    status = "PRESENT"
    return {
        "total_hours": round(total_hours, 4),
        "overtime": round_overtime_hours(ot),
        "late_mark": late,
        "status": status,
    }


def snapshot_schedule_fields(profile):
    return {
        "scheduled_shift_start": str(profile["shift_start"]),
        "scheduled_shift_end": str(profile["shift_end"]),
        "scheduled_grace_minutes": int(profile["grace_minutes"] or 0),
    }


def profile_for_attendance(conn, user_id, attendance_row=None):
    if attendance_row:
        sst = attendance_row["scheduled_shift_start"] if "scheduled_shift_start" in attendance_row.keys() else None
        set_ = attendance_row["scheduled_shift_end"] if "scheduled_shift_end" in attendance_row.keys() else None
        sgm = attendance_row["scheduled_grace_minutes"] if "scheduled_grace_minutes" in attendance_row.keys() else None
        if sst and set_ and sgm is not None:
            return {
                "shift_start": sst,
                "shift_end": set_,
                "grace_minutes": int(sgm),
            }
    p = user_profile(conn, user_id)
    if not p:
        return {"shift_start": "09:00", "shift_end": "18:00", "grace_minutes": 15}
    return {
        "shift_start": p["shift_start"],
        "shift_end": p["shift_end"],
        "grace_minutes": int(p["grace_minutes"] or 0),
    }


def late_mark_for_login(login_iso, profile):
    login = parse_dt(login_iso).astimezone(IST)
    st = datetime.combine(login.date(), time.fromisoformat(profile["shift_start"]), tzinfo=IST)
    return int(login > st + timedelta(minutes=int(profile["grace_minutes"] or 0)))


def execute_attendance_session_action(conn, user_id, purpose, action_iso, login_method=None):
    if purpose == "login":
        day = now_ist().date().isoformat()
        method = str(login_method or "UNKNOWN").strip().upper()
        if method not in {"QR", "OTP"}:
            method = "UNKNOWN"
        profile = profile_for_attendance(conn, user_id)
        schedule = snapshot_schedule_fields(profile)
        late_mark = late_mark_for_login(action_iso, profile)
        open_row = conn.execute("SELECT id FROM attendance WHERE user_id=? AND login_time IS NOT NULL AND logout_time IS NULL ORDER BY id DESC LIMIT 1", (user_id,)).fetchone()
        if open_row:
            return {"message": "User already logged in"}, 400

        today = conn.execute("SELECT id,login_time,logout_time FROM attendance WHERE user_id=? AND attendance_date=?", (user_id, day)).fetchone()
        if today and today["login_time"] and today["logout_time"]:
            return {"message": "Attendance already captured for today"}, 400

        if today and not today["login_time"]:
            conn.execute(
                """
                UPDATE attendance
                SET login_time=?,login_method=?,total_hours=NULL,overtime=NULL,late_mark=?,status='PRESENT',system_logout=0,
                    scheduled_shift_start=?,scheduled_shift_end=?,scheduled_grace_minutes=?,updated_at=?
                WHERE id=?
                """,
                (action_iso, method, late_mark, schedule["scheduled_shift_start"], schedule["scheduled_shift_end"], schedule["scheduled_grace_minutes"], action_iso, today["id"]),
            )
        else:
            conn.execute(
                """
                INSERT INTO attendance (
                    user_id,attendance_date,login_time,login_method,late_mark,break_taken,status,
                    scheduled_shift_start,scheduled_shift_end,scheduled_grace_minutes,created_at,updated_at
                )
                VALUES (?,?,?,?,?,1,'PRESENT',?,?,?,?,?)
                """,
                (user_id, day, action_iso, method, late_mark, schedule["scheduled_shift_start"], schedule["scheduled_shift_end"], schedule["scheduled_grace_minutes"], action_iso, action_iso),
            )
        return {"message": "Login Recorded"}, 200

    if purpose == "logout":
        rec = conn.execute("SELECT * FROM attendance WHERE user_id=? AND login_time IS NOT NULL AND logout_time IS NULL ORDER BY id DESC LIMIT 1", (user_id,)).fetchone()
        if not rec:
            return {"message": "No login found"}, 400
        profile = profile_for_attendance(conn, user_id, rec)
        m = calc_metrics(rec["login_time"], action_iso, profile)
        conn.execute("UPDATE attendance SET logout_time=?,total_hours=?,overtime=?,late_mark=?,status=?,updated_at=? WHERE id=?", (action_iso, m["total_hours"], m["overtime"], m["late_mark"], m["status"], action_iso, rec["id"]))
        return {"message": "Logout Recorded", "metrics": m}, 200

    return {"message": "Invalid or expired credential"}, 400


def rebuild_schema_if_needed(conn):
    conn.execute("DROP INDEX IF EXISTS idx_edit_req_status_created")
    conn.execute("DROP TABLE IF EXISTS attendance_edit_requests")
    conn.execute("DROP TABLE IF EXISTS attendance_audit_log")

    if table_exists(conn, "employee_categories") and has_col(conn, "employee_categories", "half_day_hours"):
        conn.execute("ALTER TABLE employee_categories RENAME TO employee_categories_old")
        conn.execute(
            """
            CREATE TABLE employee_categories(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                required_hours REAL NOT NULL
            )
            """
        )
        conn.execute(
            """
            INSERT INTO employee_categories (id,name,required_hours)
            SELECT id,name,required_hours FROM employee_categories_old
            """
        )
        conn.execute("DROP TABLE employee_categories_old")

    if table_exists(conn, "shifts") and has_col(conn, "shifts", "half_day_threshold"):
        conn.execute("ALTER TABLE shifts RENAME TO shifts_old")
        conn.execute(
            """
            CREATE TABLE shifts(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                grace_minutes INTEGER NOT NULL DEFAULT 15
            )
            """
        )
        conn.execute(
            """
            INSERT INTO shifts (id,name,start_time,end_time,grace_minutes)
            SELECT id,name,start_time,end_time,COALESCE(grace_minutes,15) FROM shifts_old
            """
        )
        conn.execute("DROP TABLE shifts_old")

    if table_exists(conn, "attendance") and (has_col(conn, "attendance", "early_leaving") or has_col(conn, "attendance", "half_day")):
        conn.execute("ALTER TABLE attendance RENAME TO attendance_old")
        conn.execute(
            """
            CREATE TABLE attendance(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                attendance_date TEXT NOT NULL,
                login_time TEXT,
                login_method TEXT,
                logout_time TEXT,
                total_hours REAL,
                overtime REAL,
                late_mark INTEGER NOT NULL DEFAULT 0,
                break_taken INTEGER NOT NULL DEFAULT 1,
                scheduled_shift_start TEXT,
                scheduled_shift_end TEXT,
                scheduled_grace_minutes INTEGER,
                status TEXT NOT NULL DEFAULT 'PRESENT',
                system_logout INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                UNIQUE(user_id,attendance_date)
            )
            """
        )
        conn.execute(
            """
            INSERT INTO attendance (
                id,user_id,attendance_date,login_time,login_method,logout_time,total_hours,overtime,late_mark,break_taken,scheduled_shift_start,scheduled_shift_end,scheduled_grace_minutes,status,system_logout,created_at,updated_at
            )
            WITH normalized AS (
                SELECT *,
                       COALESCE(attendance_date,substr(login_time,1,10),substr(created_at,1,10),?) normalized_date
                FROM attendance_old
            ),
            ranked AS (
                SELECT *,
                       ROW_NUMBER() OVER (PARTITION BY user_id,normalized_date ORDER BY id DESC) rn
                FROM normalized
            )
            SELECT
                id,user_id,normalized_date,login_time,NULL,logout_time,total_hours,overtime,
                COALESCE(late_mark,0),
                COALESCE(break_taken,1),
                NULL,NULL,NULL,
                CASE WHEN status='ABSENT' THEN 'ABSENT' ELSE 'PRESENT' END,
                COALESCE(system_logout,0),
                COALESCE(created_at,login_time,?),
                COALESCE(updated_at,logout_time,login_time,?)
            FROM ranked
            WHERE rn=1
            """,
            (now_ist().date().isoformat(), now_iso(), now_iso()),
        )
        conn.execute("DROP TABLE attendance_old")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_attendance_user_date ON attendance(user_id,attendance_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_attendance_status_date ON attendance(status,attendance_date)")

    # Older databases may have break_taken defaulted to 0. Rebuild once so new
    # attendance rows default to break_taken=1.
    if table_exists(conn, "attendance") and has_col(conn, "attendance", "break_taken"):
        break_default = col_default(conn, "attendance", "break_taken")
        if break_default != "1":
            conn.execute("ALTER TABLE attendance RENAME TO attendance_old_break_default")
            conn.execute(
                """
                CREATE TABLE attendance(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    attendance_date TEXT NOT NULL,
                    login_time TEXT,
                    login_method TEXT,
                    logout_time TEXT,
                    total_hours REAL,
                    overtime REAL,
                    late_mark INTEGER NOT NULL DEFAULT 0,
                    break_taken INTEGER NOT NULL DEFAULT 1,
                    scheduled_shift_start TEXT,
                    scheduled_shift_end TEXT,
                    scheduled_grace_minutes INTEGER,
                    status TEXT NOT NULL DEFAULT 'PRESENT',
                    system_logout INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT,
                    UNIQUE(user_id,attendance_date)
                )
                """
            )
            conn.execute(
                """
                INSERT INTO attendance (
                    id,user_id,attendance_date,login_time,login_method,logout_time,total_hours,overtime,late_mark,break_taken,scheduled_shift_start,scheduled_shift_end,scheduled_grace_minutes,status,system_logout,created_at,updated_at
                )
                WITH normalized AS (
                    SELECT *,
                           COALESCE(attendance_date,substr(login_time,1,10),substr(created_at,1,10),?) normalized_date
                    FROM attendance_old_break_default
                ),
                ranked AS (
                    SELECT *,
                           ROW_NUMBER() OVER (PARTITION BY user_id,normalized_date ORDER BY id DESC) rn
                    FROM normalized
                )
                SELECT
                    id,user_id,normalized_date,login_time,NULL,logout_time,total_hours,overtime,
                    COALESCE(late_mark,0),
                    COALESCE(break_taken,1),
                    NULL,NULL,NULL,
                    CASE WHEN status='ABSENT' THEN 'ABSENT' ELSE 'PRESENT' END,
                    COALESCE(system_logout,0),
                    COALESCE(created_at,login_time,?),
                    COALESCE(updated_at,logout_time,login_time,?)
                FROM ranked
                WHERE rn=1
                """,
                (now_ist().date().isoformat(), now_iso(), now_iso()),
            )
            conn.execute("DROP TABLE attendance_old_break_default")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_attendance_user_date ON attendance(user_id,attendance_date)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_attendance_status_date ON attendance(status,attendance_date)")

    conn.execute("DROP INDEX IF EXISTS idx_edit_req_status_created")


def init_db():
    with conn_db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS employee_categories(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE NOT NULL,required_hours REAL NOT NULL);
            CREATE TABLE IF NOT EXISTS shifts(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE NOT NULL,start_time TEXT NOT NULL,end_time TEXT NOT NULL,grace_minutes INTEGER NOT NULL DEFAULT 15);
            CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY,name TEXT,role TEXT NOT NULL DEFAULT 'EMPLOYEE',employee_code TEXT,pin_hash TEXT,category_id INTEGER,shift_id INTEGER,category_hours INTEGER,active INTEGER NOT NULL DEFAULT 1,created_at TEXT);
            CREATE TABLE IF NOT EXISTS attendance(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER NOT NULL,attendance_date TEXT NOT NULL,login_time TEXT,login_method TEXT,logout_time TEXT,total_hours REAL,overtime REAL,late_mark INTEGER NOT NULL DEFAULT 0,break_taken INTEGER NOT NULL DEFAULT 1,scheduled_shift_start TEXT,scheduled_shift_end TEXT,scheduled_grace_minutes INTEGER,status TEXT NOT NULL DEFAULT 'PRESENT',system_logout INTEGER NOT NULL DEFAULT 0,created_at TEXT NOT NULL,updated_at TEXT,UNIQUE(user_id,attendance_date));
            CREATE TABLE IF NOT EXISTS qr_sessions(id TEXT PRIMARY KEY,user_id INTEGER NOT NULL,purpose TEXT NOT NULL,expires_at INTEGER NOT NULL,used INTEGER NOT NULL DEFAULT 0,otp_hash TEXT,otp_failed_attempts INTEGER NOT NULL DEFAULT 0,created_at TEXT NOT NULL);
            CREATE INDEX IF NOT EXISTS idx_attendance_user_date ON attendance(user_id,attendance_date);
            CREATE INDEX IF NOT EXISTS idx_attendance_status_date ON attendance(status,attendance_date);
            CREATE INDEX IF NOT EXISTS idx_qr_sessions_user_active ON qr_sessions(user_id,used,expires_at,created_at);
            """
        )
        rebuild_schema_if_needed(conn)
        ensure_col(conn, "users", "role", "TEXT NOT NULL DEFAULT 'EMPLOYEE'")
        ensure_col(conn, "users", "employee_code", "TEXT")
        ensure_col(conn, "users", "pin_hash", "TEXT")
        ensure_col(conn, "users", "pin_plain", "TEXT")
        ensure_col(conn, "users", "category_id", "INTEGER")
        ensure_col(conn, "users", "shift_id", "INTEGER")
        ensure_col(conn, "users", "active", "INTEGER NOT NULL DEFAULT 1")
        ensure_col(conn, "users", "created_at", "TEXT")
        ensure_col(conn, "attendance", "attendance_date", "TEXT")
        ensure_col(conn, "attendance", "login_method", "TEXT")
        ensure_col(conn, "attendance", "late_mark", "INTEGER NOT NULL DEFAULT 0")
        ensure_col(conn, "attendance", "break_taken", "INTEGER NOT NULL DEFAULT 1")
        ensure_col(conn, "attendance", "scheduled_shift_start", "TEXT")
        ensure_col(conn, "attendance", "scheduled_shift_end", "TEXT")
        ensure_col(conn, "attendance", "scheduled_grace_minutes", "INTEGER")
        ensure_col(conn, "attendance", "status", "TEXT NOT NULL DEFAULT 'PRESENT'")
        ensure_col(conn, "attendance", "system_logout", "INTEGER NOT NULL DEFAULT 0")
        ensure_col(conn, "attendance", "created_at", "TEXT")
        ensure_col(conn, "attendance", "updated_at", "TEXT")
        ensure_col(conn, "qr_sessions", "created_at", "TEXT")
        ensure_col(conn, "qr_sessions", "otp_hash", "TEXT")
        ensure_col(conn, "qr_sessions", "otp_failed_attempts", "INTEGER NOT NULL DEFAULT 0")
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_employee_code ON users(employee_code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_qr_sessions_user_active ON qr_sessions(user_id,used,expires_at,created_at)")
        conn.execute("INSERT OR IGNORE INTO employee_categories (id,name,required_hours) VALUES (1,'General',9)")
        conn.execute("INSERT OR IGNORE INTO shifts (id,name,start_time,end_time,grace_minutes) VALUES (1,'General Shift','09:00','18:00',15)")
        conn.execute("INSERT OR IGNORE INTO users (id,name,role,employee_code,category_id,shift_id,category_hours,active,created_at) VALUES (1,'Employee1','EMPLOYEE','EMP001',1,1,9,1,?)", (now_iso(),))
        conn.execute("INSERT OR IGNORE INTO users (id,name,role,employee_code,category_id,shift_id,category_hours,active,created_at) VALUES (999,'Admin','ADMIN','ADMIN001',1,1,9,1,?)", (now_iso(),))
        users = conn.execute("SELECT id,role,pin_hash,employee_code FROM users").fetchall()
        for u in users:
            if not u["pin_hash"]:
                pin = DEFAULT_ADMIN_PIN if u["role"] == "ADMIN" else DEFAULT_EMPLOYEE_PIN
                conn.execute("UPDATE users SET pin_hash=?,pin_plain=? WHERE id=?", (generate_password_hash(pin), pin, u["id"]))
            if not u["employee_code"]:
                prefix = "ADMIN" if u["role"] == "ADMIN" else "EMP"
                conn.execute("UPDATE users SET employee_code=? WHERE id=?", (f"{prefix}{int(u['id']):03d}", u["id"]))
        conn.execute("UPDATE attendance SET attendance_date=COALESCE(attendance_date,substr(login_time,1,10),substr(created_at,1,10),?) WHERE attendance_date IS NULL", (now_ist().date().isoformat(),))
        conn.execute("UPDATE attendance SET created_at=COALESCE(created_at,login_time,?),updated_at=COALESCE(updated_at,logout_time,login_time,?) WHERE created_at IS NULL OR updated_at IS NULL", (now_iso(), now_iso()))
        conn.execute(
            """
            UPDATE attendance
            SET scheduled_shift_start=COALESCE(scheduled_shift_start, (
                    SELECT COALESCE(s.start_time,'09:00')
                    FROM users u
                    LEFT JOIN shifts s ON s.id=u.shift_id
                    WHERE u.id=attendance.user_id
                )),
                scheduled_shift_end=COALESCE(scheduled_shift_end, (
                    SELECT COALESCE(s.end_time,'18:00')
                    FROM users u
                    LEFT JOIN shifts s ON s.id=u.shift_id
                    WHERE u.id=attendance.user_id
                )),
                scheduled_grace_minutes=COALESCE(scheduled_grace_minutes, (
                    SELECT COALESCE(s.grace_minutes,15)
                    FROM users u
                    LEFT JOIN shifts s ON s.id=u.shift_id
                    WHERE u.id=attendance.user_id
                ))
            WHERE scheduled_shift_start IS NULL OR scheduled_shift_end IS NULL OR scheduled_grace_minutes IS NULL
            """
        )
        conn.execute("UPDATE qr_sessions SET created_at=COALESCE(created_at,?) WHERE created_at IS NULL", (now_iso(),))
        conn.execute("UPDATE qr_sessions SET otp_failed_attempts=COALESCE(otp_failed_attempts,0) WHERE otp_failed_attempts IS NULL")


def auth_user():
    uid = session.get("user_id")
    if not uid:
        return None
    with conn_db() as conn:
        u = conn.execute("SELECT id,name,role,employee_code,active FROM users WHERE id=?", (uid,)).fetchone()
    if not u or int(u["active"] or 0) != 1:
        return None
    return u


def html_guard(role=None):
    def d(fn):
        @wraps(fn)
        def w(*a, **k):
            u = auth_user()
            if not u:
                return redirect(url_for("login_page"))
            if role and u["role"] != role:
                return redirect(url_for("dashboard"))
            if not role_bypasses_office_check(u["role"]):
                chk = office_check()
                if chk:
                    return chk
            return fn(*a, **k)

        return w

    return d


def api_guard(role=None):
    def d(fn):
        @wraps(fn)
        def w(*a, **k):
            u = auth_user()
            if not u:
                return jsonify({"message": "Authentication required"}), 401
            if role and u["role"] != role:
                return jsonify({"message": "Forbidden"}), 403
            if not role_bypasses_office_check(u["role"]):
                chk = office_check()
                if chk:
                    return chk
            request.current_user = u
            return fn(*a, **k)

        return w

    return d


@app.get("/")
def root():
    return redirect(url_for("login_page"))


@app.get("/health")
def health():
    return jsonify({"status": "ok", "time_ist": now_iso()})


@app.get("/admin.html")
def admin_tool_page():
    return redirect(url_for("scanner_page"))


@app.get("/scanner")
def scanner_page():
    chk = office_check()
    if chk:
        return chk
    return send_from_directory(PUBLIC_DIR, "scanner.html")


@app.get("/employee.html")
@html_guard("EMPLOYEE")
def employee_tool_page():
    return send_from_directory(PUBLIC_DIR, "employee.html")


@app.get("/scanner.js")
def scanner_js():
    return send_from_directory(PUBLIC_DIR, "scanner.js")


@app.get("/login")
def login_page():
    if not office_ok() and not ALLOW_ADMIN_FROM_ANYWHERE:
        return jsonify({"message": "Access allowed only from office network"}), 403
    if auth_user():
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.get("/dashboard")
@html_guard()
def dashboard():
    u = auth_user()
    return redirect(url_for("admin_dashboard" if u["role"] == "ADMIN" else "employee_dashboard"))


@app.get("/dashboard/admin")
@html_guard("ADMIN")
def admin_dashboard():
    return render_template("admin_dashboard.html")


@app.get("/dashboard/employee")
@html_guard("EMPLOYEE")
def employee_dashboard():
    return render_template("employee_dashboard.html")


@app.post("/auth/login")
def auth_login():
    data = request.get_json(silent=True) or request.form.to_dict() or {}
    pin = str(data.get("pin", "")).strip()
    if not valid_pin(pin):
        return jsonify({"message": "PIN must be 4 digits"}), 400

    key = str(data.get("employee_code", "")).strip()
    uid = data.get("user_id")
    with conn_db() as conn:
        user = None
        if key:
            user = conn.execute("SELECT * FROM users WHERE employee_code=? AND active=1", (key,)).fetchone()
        elif uid is not None:
            try:
                user = conn.execute("SELECT * FROM users WHERE id=? AND active=1", (int(uid),)).fetchone()
            except ValueError:
                return jsonify({"message": "Invalid user_id"}), 400
        else:
            return jsonify({"message": "employee_code or user_id required"}), 400

        if not user or not user["pin_hash"] or not check_password_hash(user["pin_hash"], pin):
            return jsonify({"message": "Invalid credentials"}), 401

    if not office_ok() and not role_bypasses_office_check(user["role"]):
        return jsonify({"message": "Access allowed only from office network"}), 403

    session["user_id"] = int(user["id"])
    session["role"] = user["role"]
    session["name"] = user["name"]
    session["login_at"] = now_iso()
    return jsonify({"message": "Login successful", "redirect": "/dashboard/admin" if user["role"] == "ADMIN" else "/dashboard/employee"})


@app.post("/auth/logout")
def auth_logout():
    session.clear()
    return jsonify({"message": "Logged out"})


@app.get("/auth/me")
@api_guard()
def auth_me():
    u = request.current_user
    with conn_db() as conn:
        p = user_profile(conn, int(u["id"]))
    category_name = p["category_name"] if p else "General"
    shift_start = p["shift_start"] if p else "09:00"
    shift_end = p["shift_end"] if p else "18:00"
    return jsonify(
        {
            "user": {
                "id": int(u["id"]),
                "name": u["name"],
                "role": u["role"],
                "employee_code": u["employee_code"],
                "category_name": category_name,
                "shift_start": shift_start,
                "shift_end": shift_end,
            }
        }
    )

@app.post("/generate-qr")
@api_guard("EMPLOYEE")
def generate_qr():
    u = request.current_user
    data = request.get_json(silent=True) or {}
    purpose = str(data.get("purpose", "")).strip().lower()
    if purpose not in {"login", "logout"}:
        return jsonify({"message": "Invalid request payload"}), 400

    user_id = int(u["id"])
    token = str(uuid.uuid4())
    otp_code = generate_otp_code()
    nowms = epoch_ms()
    expires = nowms + QR_TTL_SECONDS * 1000
    with conn_db() as conn:
        if not user_profile(conn, user_id):
            return jsonify({"message": "User not found or inactive"}), 404
        conn.execute("UPDATE qr_sessions SET used=1 WHERE user_id=? AND used=0 AND expires_at>=?", (user_id, nowms))
        conn.execute(
            "INSERT INTO qr_sessions (id,user_id,purpose,expires_at,used,otp_hash,otp_failed_attempts,created_at) VALUES (?,?,?,?,0,?,?,?)",
            (token, user_id, purpose, expires, otp_digest(otp_code, token), 0, now_iso()),
        )
    payload = {"user_id": user_id, "session_token": token}
    qr_text = json.dumps(payload)
    img = qrcode.make(qr_text)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    qr = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")
    return jsonify(
        {
            "qr": qr,
            "session_id": token,
            "session_token": token,
            "otp_code": otp_code,
            "expires_in_seconds": QR_TTL_SECONDS,
            "employee_code": u["employee_code"],
            "employee_name": u["name"],
        }
    )


@app.post("/scan")
def scan_qr():
    chk = office_check()
    if chk:
        return chk
    data = request.get_json(silent=True) or {}
    token = str((data.get("session_token") or data.get("session_id") or "")).strip()
    employee_code = str(data.get("employee_code", "")).strip()
    otp_code = str(data.get("otp_code", "")).strip()
    qr_mode = bool(token)
    otp_mode = bool(employee_code or otp_code)

    if qr_mode and otp_mode:
        return jsonify({"message": "Provide either QR session token or employee_code + otp_code"}), 400
    if not qr_mode and not otp_mode:
        return jsonify({"message": "Provide either QR session token or employee_code + otp_code"}), 400
    if otp_mode:
        if not employee_code or not otp_code:
            return jsonify({"message": "employee_code and otp_code are required"}), 400
        if not otp_code.isdigit() or len(otp_code) != OTP_LENGTH:
            return jsonify({"message": f"otp_code must be {OTP_LENGTH} digits"}), 400

    nowms = epoch_ms()
    niso = now_iso()
    with conn_db() as conn:
        if qr_mode:
            qr = conn.execute("SELECT * FROM qr_sessions WHERE id=?", (token,)).fetchone()
            if not qr or int(qr["used"]) == 1 or int(qr["expires_at"]) < nowms:
                return jsonify({"message": "Invalid or expired credential"}), 400
            conn.execute("UPDATE qr_sessions SET used=1 WHERE id=?", (token,))
            payload, status = execute_attendance_session_action(conn, int(qr["user_id"]), str(qr["purpose"]).lower(), niso, "QR")
            return jsonify(payload), status

        user = conn.execute("SELECT id FROM users WHERE UPPER(employee_code)=UPPER(?) AND active=1", (employee_code,)).fetchone()
        if not user:
            return jsonify({"message": "Invalid employee code or OTP"}), 400
        uid = int(user["id"])
        qr = conn.execute(
            """
            SELECT * FROM qr_sessions
            WHERE user_id=? AND used=0 AND expires_at>=?
            ORDER BY created_at DESC LIMIT 1
            """,
            (uid, nowms),
        ).fetchone()
        if not qr or not qr["otp_hash"]:
            return jsonify({"message": "Invalid employee code or OTP"}), 400

        attempts = int(qr["otp_failed_attempts"] or 0)
        if attempts >= OTP_MAX_ATTEMPTS:
            conn.execute("UPDATE qr_sessions SET used=1 WHERE id=?", (qr["id"],))
            return jsonify({"message": "OTP locked. Generate a new QR/OTP."}), 400

        expected = str(qr["otp_hash"])
        actual = otp_digest(otp_code, str(qr["id"]))
        if not hmac.compare_digest(actual, expected):
            new_attempts = attempts + 1
            locked = new_attempts >= OTP_MAX_ATTEMPTS
            conn.execute("UPDATE qr_sessions SET otp_failed_attempts=?,used=? WHERE id=?", (new_attempts, 1 if locked else 0, qr["id"]))
            if locked:
                return jsonify({"message": "OTP locked after too many failed attempts. Generate a new QR/OTP."}), 400
            remaining = OTP_MAX_ATTEMPTS - new_attempts
            return jsonify({"message": f"Invalid OTP. {remaining} attempt(s) remaining."}), 400

        conn.execute("UPDATE qr_sessions SET used=1 WHERE id=?", (qr["id"],))
        payload, status = execute_attendance_session_action(conn, uid, str(qr["purpose"]).lower(), niso, "OTP")
        return jsonify(payload), status


@app.post("/admin/run-midnight-close")
def midnight_close():
    chk = office_check()
    if chk:
        return chk
    data = request.get_json(silent=True) or {}
    d = data.get("date")
    try:
        target = date.fromisoformat(d) if d else now_ist().date() - timedelta(days=1)
    except ValueError:
        return jsonify({"message": "Invalid date format. Use YYYY-MM-DD"}), 400

    absent = 0
    auto_logout = 0
    with conn_db() as conn:
        users = conn.execute("SELECT id FROM users WHERE active=1").fetchall()
        for u in users:
            uid = int(u["id"])
            row = conn.execute("SELECT * FROM attendance WHERE user_id=? AND attendance_date=?", (uid, target.isoformat())).fetchone()
            if not row:
                conn.execute("INSERT INTO attendance (user_id,attendance_date,break_taken,status,created_at,updated_at) VALUES (?, ?, 0, 'ABSENT', ?, ?)", (uid, target.isoformat(), now_iso(), now_iso()))
                absent += 1
                continue
            if row["login_time"] and not row["logout_time"]:
                p = profile_for_attendance(conn, uid, row)
                st = datetime.combine(target, time.fromisoformat(p["shift_start"]), tzinfo=IST)
                et = datetime.combine(target, time.fromisoformat(p["shift_end"]), tzinfo=IST)
                if et <= st:
                    et += timedelta(days=1)
                auto_t = et.isoformat()
                m = calc_metrics(row["login_time"], auto_t, p)
                conn.execute("UPDATE attendance SET logout_time=?,total_hours=?,overtime=?,late_mark=?,status=?,system_logout=1,updated_at=? WHERE id=?", (auto_t, m["total_hours"], m["overtime"], m["late_mark"], m["status"], now_iso(), row["id"]))
                auto_logout += 1

    return jsonify({"message": "Midnight close completed", "attendance_date": target.isoformat(), "absent_marked": absent, "system_logout_done": auto_logout})


@app.get("/admin/backup-db")
def backup_db():
    ok, message = backup_db_to_github()
    if not ok:
        app.logger.warning("Backup request failed: %s", message)
        return {"message": message}, 500
    return {"message": message}


def parse_filters():
    try:
        dfrom, dto = resolve_date_range(request.args.get("from"), request.args.get("to"))
    except Exception as e:
        raise ValueError(str(e))
    status = request.args.get("status")
    if status:
        status = status.strip().upper()
        if status not in {"PRESENT", "ABSENT"}:
            raise ValueError("Invalid status")
    uid = int(request.args.get("user_id")) if request.args.get("user_id") else None
    sid = int(request.args.get("shift_id")) if request.args.get("shift_id") else None
    cid = int(request.args.get("category_id")) if request.args.get("category_id") else None
    return dfrom, dto, status, uid, sid, cid


def attendance_where(dfrom, dto, status=None, uid=None, sid=None, cid=None):
    cond = ["a.attendance_date BETWEEN ? AND ?"]
    params = [dfrom.isoformat(), dto.isoformat()]
    if status:
        cond.append("a.status=?")
        params.append(status)
    if uid is not None:
        cond.append("a.user_id=?")
        params.append(uid)
    if sid is not None:
        cond.append("u.shift_id=?")
        params.append(sid)
    if cid is not None:
        cond.append("u.category_id=?")
        params.append(cid)
    return " AND ".join(cond), params


@app.get("/api/admin/attendance")
@api_guard("ADMIN")
def admin_attendance():
    try:
        dfrom, dto, status, uid, sid, cid = parse_filters()
        page = max(1, int(request.args.get("page", 1)))
        page_size = max(1, min(200, int(request.args.get("page_size", 25))))
    except Exception as e:
        return jsonify({"message": str(e)}), 400
    offset = (page - 1) * page_size

    where, params = attendance_where(dfrom, dto, status, uid, sid, cid)
    with conn_db() as conn:
        total = conn.execute(f"SELECT COUNT(*) cnt FROM attendance a JOIN users u ON u.id=a.user_id WHERE {where}", tuple(params)).fetchone()["cnt"]
        rows = conn.execute(
            f"""
            SELECT a.*,u.name employee_name,u.employee_code,COALESCE(s.name,'General Shift') shift_name,COALESCE(c.name,'General') category_name,
                   CASE
                     WHEN a.login_time IS NULL OR a.logout_time IS NULL THEN NULL
                     ELSE ROUND(MAX(0, COALESCE(c.required_hours,9) - COALESCE(a.total_hours,0)), 4)
                   END early_logout_hours
            FROM attendance a JOIN users u ON u.id=a.user_id
            LEFT JOIN shifts s ON s.id=u.shift_id LEFT JOIN employee_categories c ON c.id=u.category_id
            WHERE {where} ORDER BY a.attendance_date DESC,a.id DESC LIMIT ? OFFSET ?
            """,
            tuple(params + [page_size, offset]),
        ).fetchall()
    return jsonify({"items": to_items(rows), "page": page, "page_size": page_size, "total": total})


@app.get("/api/admin/summary")
@api_guard("ADMIN")
def admin_summary():
    try:
        dfrom, dto = resolve_date_range(request.args.get("from"), request.args.get("to"))
    except Exception as e:
        return jsonify({"message": str(e)}), 400
    with conn_db() as conn:
        row = conn.execute(
            """
            SELECT SUM(CASE WHEN login_time IS NOT NULL THEN 1 ELSE 0 END) total_days_worked,
            SUM(CASE WHEN status='ABSENT' THEN 1 ELSE 0 END) absent_count,
            SUM(CASE WHEN late_mark=1 THEN 1 ELSE 0 END) late_count,
            SUM(COALESCE(total_hours,0)) total_hours,
            SUM(COALESCE(overtime,0)) overtime_hours
            FROM attendance WHERE attendance_date BETWEEN ? AND ?
            """,
            (dfrom.isoformat(), dto.isoformat()),
        ).fetchone()
    return jsonify(dict(row))


@app.get("/api/admin/employee-summary")
@api_guard("ADMIN")
def admin_employee_summary():
    code = str(request.args.get("employee_code", "")).strip()
    if not code:
        return jsonify({"message": "employee_code is required"}), 400
    try:
        dfrom, dto = resolve_date_range(request.args.get("from"), request.args.get("to"))
    except Exception as e:
        return jsonify({"message": str(e)}), 400

    with conn_db() as conn:
        employee = conn.execute(
            """
            SELECT u.id,u.name,u.employee_code
            FROM users u
            WHERE u.employee_code=?
            """,
            (code,),
        ).fetchone()
        if not employee:
            return jsonify({"message": "Employee not found"}), 404

        summary = conn.execute(
            """
            SELECT
            SUM(CASE WHEN status='ABSENT' THEN 1 ELSE 0 END) absent_count,
            SUM(CASE WHEN login_time IS NOT NULL THEN 1 ELSE 0 END) total_days_worked,
            SUM(CASE WHEN late_mark=1 THEN 1 ELSE 0 END) late_count,
            SUM(COALESCE(total_hours,0)) total_hours,
            SUM(COALESCE(overtime,0)) overtime_hours
            FROM attendance WHERE user_id=? AND attendance_date BETWEEN ? AND ?
            """,
            (employee["id"], dfrom.isoformat(), dto.isoformat()),
        ).fetchone()
        rows = conn.execute(
            """
            SELECT a.id,a.attendance_date,a.login_time,a.login_method,a.logout_time,a.total_hours,a.overtime,a.late_mark,a.break_taken,a.status,
                   CASE
                     WHEN a.login_time IS NULL OR a.logout_time IS NULL THEN NULL
                     ELSE ROUND(MAX(0, COALESCE(c.required_hours,9) - COALESCE(a.total_hours,0)), 4)
                   END early_logout_hours
            FROM attendance a
            JOIN users u ON u.id=a.user_id
            LEFT JOIN employee_categories c ON c.id=u.category_id
            WHERE a.user_id=? AND a.attendance_date BETWEEN ? AND ?
            ORDER BY a.attendance_date DESC,a.id DESC
            """,
            (employee["id"], dfrom.isoformat(), dto.isoformat()),
        ).fetchall()

    return jsonify({"employee": dict(employee), "summary": dict(summary), "attendance": to_items(rows)})


@app.get("/api/admin/employee-summary.xlsx")
@api_guard("ADMIN")
def admin_employee_summary_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({"message": "openpyxl is required for export"}), 500

    code = str(request.args.get("employee_code", "")).strip()
    if not code:
        return jsonify({"message": "employee_code is required"}), 400
    try:
        dfrom, dto = resolve_date_range(request.args.get("from"), request.args.get("to"))
    except Exception as e:
        return jsonify({"message": str(e)}), 400

    with conn_db() as conn:
        employee = conn.execute(
            """
            SELECT u.id,u.name,u.employee_code
            FROM users u
            WHERE u.employee_code=?
            """,
            (code,),
        ).fetchone()
        if not employee:
            return jsonify({"message": "Employee not found"}), 404

        summary = conn.execute(
            """
            SELECT
            SUM(CASE WHEN status='ABSENT' THEN 1 ELSE 0 END) absent_count,
            SUM(CASE WHEN login_time IS NOT NULL THEN 1 ELSE 0 END) total_days_worked,
            SUM(CASE WHEN late_mark=1 THEN 1 ELSE 0 END) late_count,
            SUM(COALESCE(total_hours,0)) total_hours,
            SUM(COALESCE(overtime,0)) overtime_hours
            FROM attendance WHERE user_id=? AND attendance_date BETWEEN ? AND ?
            """,
            (employee["id"], dfrom.isoformat(), dto.isoformat()),
        ).fetchone()
        rows = conn.execute(
            """
            SELECT attendance_date,login_time,logout_time,total_hours,overtime,late_mark,break_taken,status
            FROM attendance WHERE user_id=? AND attendance_date BETWEEN ? AND ?
            ORDER BY attendance_date DESC,id DESC
            """,
            (employee["id"], dfrom.isoformat(), dto.isoformat()),
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Summary"
    ws.append(["Employee Name", employee["name"]])
    ws.append(["Employee Code", employee["employee_code"]])
    ws.append([])
    ws.append(["From", dfrom.isoformat(), "To", dto.isoformat()])
    ws.append([])
    ws.append(["Absent", "Total Days Worked", "Late", "Total Hours", "Overtime"])
    ws.append([summary["absent_count"], summary["total_days_worked"], summary["late_count"], summary["total_hours"], summary["overtime_hours"]])
    ws.append([])
    ws.append(["Date", "Login", "Logout", "Hours", "OT", "Late", "Break", "Status"])
    for r in rows:
        ws.append([r["attendance_date"], r["login_time"], r["logout_time"], r["total_hours"], r["overtime"], r["late_mark"], r["break_taken"], r["status"]])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"employee_summary_{code}_{dfrom.isoformat()}_{dto.isoformat()}.xlsx")


@app.post("/api/admin/attendance/edit")
@api_guard("ADMIN")
def admin_attendance_edit():
    d = request.get_json(silent=True) or {}
    code = str(d.get("employee_code", "")).strip()
    att_date = str(d.get("attendance_date", "")).strip()
    if not code or not att_date:
        return jsonify({"message": "employee_code and attendance_date are required"}), 400
    try:
        date.fromisoformat(att_date)
    except ValueError:
        return jsonify({"message": "attendance_date must be YYYY-MM-DD"}), 400

    has_login = "login_time" in d
    has_logout = "logout_time" in d
    has_break = "break_taken" in d
    has_category = "category_id" in d
    has_shift = "shift_id" in d
    if not any([has_login, has_logout, has_break, has_category, has_shift]):
        return jsonify({"message": "No fields to update"}), 400
    category_id = None
    if has_category:
        try:
            category_id = int(d.get("category_id"))
        except Exception:
            return jsonify({"message": "category_id must be an integer"}), 400
    shift_id = None
    if has_shift:
        try:
            shift_id = int(d.get("shift_id"))
        except Exception:
            return jsonify({"message": "shift_id must be an integer"}), 400

    with conn_db() as conn:
        user = conn.execute("SELECT id,employee_code,category_id,shift_id FROM users WHERE employee_code=?", (code,)).fetchone()
        if not user:
            return jsonify({"message": "Employee not found"}), 404
        if has_category:
            cat = conn.execute("SELECT id FROM employee_categories WHERE id=?", (category_id,)).fetchone()
            if not cat:
                return jsonify({"message": "Category not found"}), 404
        if has_shift:
            sh = conn.execute("SELECT id FROM shifts WHERE id=?", (shift_id,)).fetchone()
            if not sh:
                return jsonify({"message": "Shift not found"}), 404
        row = conn.execute("SELECT * FROM attendance WHERE user_id=? AND attendance_date=? ORDER BY id DESC LIMIT 1", (user["id"], att_date)).fetchone()
        if not row:
            return jsonify({"message": "Attendance not found for selected date"}), 404

        fields, params = [], []
        if has_login:
            login_time = d.get("login_time")
            if login_time:
                parse_dt(login_time)
            fields.append("login_time=?")
            params.append(login_time)
        if has_logout:
            logout_time = d.get("logout_time")
            if logout_time:
                parse_dt(logout_time)
            fields.append("logout_time=?")
            params.append(logout_time)
        if has_break:
            b = d.get("break_taken")
            if isinstance(b, bool):
                break_taken = int(b)
            elif isinstance(b, (int, float)) and int(b) in {0, 1}:
                break_taken = int(b)
            elif isinstance(b, str) and b.strip().lower() in {"0", "1", "true", "false", "yes", "no"}:
                break_taken = 1 if b.strip().lower() in {"1", "true", "yes"} else 0
            else:
                return jsonify({"message": "break_taken must be true/false"}), 400
            fields.append("break_taken=?")
            params.append(break_taken)

        fields.append("updated_at=?")
        params.append(now_iso())
        params.append(row["id"])
        conn.execute(f"UPDATE attendance SET {', '.join(fields)} WHERE id=?", tuple(params))

        user_fields, user_params = [], []
        if has_category:
            user_fields.append("category_id=?")
            user_params.append(category_id)
        if has_shift:
            user_fields.append("shift_id=?")
            user_params.append(shift_id)
        if user_fields:
            user_params.append(user["id"])
            conn.execute(f"UPDATE users SET {', '.join(user_fields)} WHERE id=?", tuple(user_params))

        current_attendance = conn.execute("SELECT * FROM attendance WHERE id=?", (row["id"],)).fetchone()
        if has_shift or (current_attendance and (not current_attendance["scheduled_shift_start"] or not current_attendance["scheduled_shift_end"] or current_attendance["scheduled_grace_minutes"] is None)):
            schedule_profile = profile_for_attendance(conn, int(user["id"]))
            schedule = snapshot_schedule_fields(schedule_profile)
            conn.execute(
                "UPDATE attendance SET scheduled_shift_start=?,scheduled_shift_end=?,scheduled_grace_minutes=?,updated_at=? WHERE id=?",
                (schedule["scheduled_shift_start"], schedule["scheduled_shift_end"], schedule["scheduled_grace_minutes"], now_iso(), row["id"]),
            )

        latest = conn.execute("SELECT * FROM attendance WHERE id=?", (row["id"],)).fetchone()
        if latest["login_time"] and latest["logout_time"]:
            recalc_attendance(conn, int(row["id"]))
        else:
            conn.execute(
                "UPDATE attendance SET total_hours=NULL,overtime=NULL,late_mark=0,status='PRESENT',updated_at=? WHERE id=?",
                (now_iso(), row["id"]),
            )
        new = dict(conn.execute("SELECT * FROM attendance WHERE id=?", (row["id"],)).fetchone())
        conn.commit()
    return jsonify({"message": "Attendance updated", "item": new})


@app.get("/api/employee/my-attendance")
@api_guard("EMPLOYEE")
def my_attendance():
    u = request.current_user
    try:
        dfrom, dto = resolve_date_range(request.args.get("from"), request.args.get("to"))
        page = max(1, int(request.args.get("page", 1)))
        page_size = max(1, min(200, int(request.args.get("page_size", 25))))
    except Exception as e:
        return jsonify({"message": str(e)}), 400
    offset = (page - 1) * page_size
    with conn_db() as conn:
        total = conn.execute("SELECT COUNT(*) cnt FROM attendance WHERE user_id=? AND attendance_date BETWEEN ? AND ?", (u["id"], dfrom.isoformat(), dto.isoformat())).fetchone()["cnt"]
        rows = conn.execute(
            """
            SELECT a.*,
                   CASE
                     WHEN a.login_time IS NULL OR a.logout_time IS NULL THEN NULL
                     ELSE ROUND(MAX(0, COALESCE(c.required_hours,9) - COALESCE(a.total_hours,0)), 4)
                   END early_logout_hours
            FROM attendance a
            JOIN users u ON u.id=a.user_id
            LEFT JOIN employee_categories c ON c.id=u.category_id
            WHERE a.user_id=? AND a.attendance_date BETWEEN ? AND ?
            ORDER BY a.attendance_date DESC,a.id DESC
            LIMIT ? OFFSET ?
            """,
            (u["id"], dfrom.isoformat(), dto.isoformat(), page_size, offset),
        ).fetchall()
    return jsonify({"items": to_items(rows), "page": page, "page_size": page_size, "total": total})


@app.get("/api/employee/my-summary")
@api_guard("EMPLOYEE")
def my_summary():
    u = request.current_user
    try:
        dfrom, dto = resolve_date_range(request.args.get("from"), request.args.get("to"))
    except Exception as e:
        return jsonify({"message": str(e)}), 400
    with conn_db() as conn:
        row = conn.execute(
            """
            SELECT COUNT(*) total_days,
            SUM(CASE WHEN status='PRESENT' THEN 1 ELSE 0 END) present_count,
            SUM(CASE WHEN status='ABSENT' THEN 1 ELSE 0 END) absent_count,
            SUM(CASE WHEN late_mark=1 THEN 1 ELSE 0 END) late_count,
            SUM(CASE WHEN break_taken=1 THEN 1 ELSE 0 END) break_taken_days,
            SUM(COALESCE(total_hours,0)) total_hours,
            SUM(COALESCE(overtime,0)) overtime_hours
            FROM attendance WHERE user_id=? AND attendance_date BETWEEN ? AND ?
            """,
            (u["id"], dfrom.isoformat(), dto.isoformat()),
        ).fetchone()
    return jsonify(dict(row))


@app.get("/api/employee/today-break")
@api_guard("EMPLOYEE")
def employee_today_break():
    u = request.current_user
    effective_day = break_effective_date_ist().isoformat()
    with conn_db() as conn:
        row = conn.execute("SELECT id,attendance_date,break_taken FROM attendance WHERE user_id=? AND attendance_date=? ORDER BY id DESC LIMIT 1", (u["id"], effective_day)).fetchone()
    if not row:
        return jsonify({"attendance_date": effective_day, "break_taken": None, "can_update": False})
    return jsonify({"attendance_date": effective_day, "break_taken": int(row["break_taken"] or 0), "can_update": True})


@app.post("/api/employee/today-break")
@api_guard("EMPLOYEE")
def employee_today_break_update():
    u = request.current_user
    d = request.get_json(silent=True) or {}
    raw = d.get("break_taken")
    if isinstance(raw, bool):
        break_taken = int(raw)
    elif isinstance(raw, (int, float)) and int(raw) in {0, 1}:
        break_taken = int(raw)
    elif isinstance(raw, str) and raw.strip().lower() in {"0", "1", "true", "false", "yes", "no"}:
        break_taken = 1 if raw.strip().lower() in {"1", "true", "yes"} else 0
    else:
        return jsonify({"message": "break_taken must be true/false"}), 400

    effective_day = break_effective_date_ist().isoformat()
    with conn_db() as conn:
        row = conn.execute("SELECT * FROM attendance WHERE user_id=? AND attendance_date=? ORDER BY id DESC LIMIT 1", (u["id"], effective_day)).fetchone()
        if not row:
            return jsonify({"message": f"No attendance record found for break date {effective_day} (4:00 AM cutoff)"}), 400
        conn.execute("UPDATE attendance SET break_taken=?,updated_at=? WHERE id=?", (break_taken, now_iso(), row["id"]))
        conn.commit()
    return jsonify({"message": "Break status updated", "attendance_date": effective_day, "break_taken": break_taken})


@app.get("/api/admin/users")
@api_guard("ADMIN")
def users_list():
    with conn_db() as conn:
        rows = conn.execute("SELECT u.id,u.name,u.role,u.employee_code,u.pin_plain,u.active,u.created_at,u.category_id,u.shift_id,COALESCE(c.name,'General') category_name,COALESCE(s.name,'General Shift') shift_name FROM users u LEFT JOIN employee_categories c ON c.id=u.category_id LEFT JOIN shifts s ON s.id=u.shift_id ORDER BY u.id").fetchall()
    items = [dict(r) for r in rows]
    for i in items:
        i["pin_set"] = True
    return jsonify({"items": items})


@app.post("/api/admin/users")
@api_guard("ADMIN")
def users_create():
    d = request.get_json(silent=True) or {}
    name = str(d.get("name", "")).strip()
    role = str(d.get("role", "EMPLOYEE")).strip().upper()
    code = str(d.get("employee_code", "")).strip()
    pin = str(d.get("pin", "")).strip()
    if not name or role not in {"ADMIN", "EMPLOYEE"} or not code or not valid_pin(pin):
        return jsonify({"message": "Invalid user payload"}), 400
    cid = int(d.get("category_id", 1))
    sid = int(d.get("shift_id", 1))
    active = int(d.get("active", 1))
    with conn_db() as conn:
        try:
            cur = conn.execute("INSERT INTO users (name,role,employee_code,pin_hash,pin_plain,category_id,shift_id,category_hours,active,created_at) VALUES (?,?,?,?,?,?,?,9,?,?)", (name, role, code, generate_password_hash(pin), pin, cid, sid, active, now_iso()))
            conn.commit()
        except sqlite3.IntegrityError as e:
            return jsonify({"message": str(e)}), 400
    return jsonify({"message": "User created", "id": int(cur.lastrowid)})


@app.put("/api/admin/users/<int:user_id>")
@api_guard("ADMIN")
def users_update(user_id):
    d = request.get_json(silent=True) or {}
    fields, params = [], []
    def put(name, val):
        fields.append(f"{name}=?"); params.append(val)
    if "name" in d: put("name", str(d.get("name", "")).strip())
    if "role" in d:
        role = str(d.get("role", "")).strip().upper()
        if role not in {"ADMIN", "EMPLOYEE"}: return jsonify({"message": "Invalid role"}), 400
        put("role", role)
    if "employee_code" in d: put("employee_code", str(d.get("employee_code", "")).strip())
    if "category_id" in d: put("category_id", int(d.get("category_id")))
    if "shift_id" in d: put("shift_id", int(d.get("shift_id")))
    if "active" in d: put("active", int(d.get("active")))
    if "pin" in d:
        pin = str(d.get("pin", "")).strip()
        if not valid_pin(pin): return jsonify({"message": "PIN must be 4 digits"}), 400
        put("pin_hash", generate_password_hash(pin))
        put("pin_plain", pin)
    if not fields: return jsonify({"message": "No fields to update"}), 400
    params.append(user_id)
    with conn_db() as conn:
        try:
            conn.execute(f"UPDATE users SET {', '.join(fields)} WHERE id=?", tuple(params))
            conn.commit()
        except sqlite3.IntegrityError as e:
            return jsonify({"message": str(e)}), 400
    return jsonify({"message": "User updated"})


@app.delete("/api/admin/users/<int:user_id>")
@api_guard("ADMIN")
def users_delete(user_id):
    current = request.current_user
    with conn_db() as conn:
        user = conn.execute("SELECT id,role FROM users WHERE id=?", (user_id,)).fetchone()
        if not user:
            return jsonify({"message": "User not found"}), 404
        if int(user["id"]) == int(current["id"]):
            return jsonify({"message": "You cannot delete your own account"}), 400
        if user["role"] == "ADMIN":
            active_admins_left = conn.execute("SELECT COUNT(*) cnt FROM users WHERE role='ADMIN' AND active=1 AND id<>?", (user_id,)).fetchone()["cnt"]
            if int(active_admins_left or 0) < 1:
                return jsonify({"message": "Cannot delete the last active admin"}), 400
        conn.execute("DELETE FROM attendance WHERE user_id=?", (user_id,))
        conn.execute("DELETE FROM qr_sessions WHERE user_id=?", (user_id,))
        conn.execute("DELETE FROM users WHERE id=?", (user_id,))
        conn.commit()
    return jsonify({"message": "User deleted"})


@app.get("/api/admin/categories")
@api_guard("ADMIN")
def categories_list():
    with conn_db() as conn:
        rows = conn.execute("SELECT * FROM employee_categories ORDER BY id").fetchall()
    return jsonify({"items": to_items(rows)})


@app.post("/api/admin/categories")
@api_guard("ADMIN")
def categories_create():
    d = request.get_json(silent=True) or {}
    try:
        name = str(d.get("name", "")).strip(); req = float(d.get("required_hours"))
    except Exception:
        return jsonify({"message": "Invalid category payload"}), 400
    with conn_db() as conn:
        try:
            cur = conn.execute("INSERT INTO employee_categories (name,required_hours) VALUES (?,?)", (name, req)); conn.commit()
        except sqlite3.IntegrityError as e:
            return jsonify({"message": str(e)}), 400
    return jsonify({"message": "Category created", "id": int(cur.lastrowid)})


@app.put("/api/admin/categories/<int:category_id>")
@api_guard("ADMIN")
def categories_update(category_id):
    d = request.get_json(silent=True) or {}
    fields, params = [], []
    if "name" in d: fields.append("name=?"); params.append(str(d.get("name", "")).strip())
    if "required_hours" in d: fields.append("required_hours=?"); params.append(float(d.get("required_hours")))
    if not fields: return jsonify({"message": "No fields to update"}), 400
    params.append(category_id)
    with conn_db() as conn:
        try:
            conn.execute(f"UPDATE employee_categories SET {', '.join(fields)} WHERE id=?", tuple(params)); conn.commit()
        except sqlite3.IntegrityError as e:
            return jsonify({"message": str(e)}), 400
    return jsonify({"message": "Category updated"})


@app.delete("/api/admin/categories/<int:category_id>")
@api_guard("ADMIN")
def categories_delete(category_id):
    with conn_db() as conn:
        row = conn.execute("SELECT id FROM employee_categories WHERE id=?", (category_id,)).fetchone()
        if not row:
            return jsonify({"message": "Category not found"}), 404
        fallback = conn.execute("SELECT id FROM employee_categories WHERE id<>? ORDER BY id LIMIT 1", (category_id,)).fetchone()
        if not fallback:
            return jsonify({"message": "Cannot delete the last remaining category"}), 400
        fallback_id = int(fallback["id"])
        conn.execute("UPDATE users SET category_id=? WHERE category_id=?", (fallback_id, category_id))
        conn.execute("DELETE FROM employee_categories WHERE id=?", (category_id,))
        conn.commit()
    return jsonify({"message": "Category deleted", "fallback_category_id": fallback_id})


@app.get("/api/admin/shifts")
@api_guard("ADMIN")
def shifts_list():
    with conn_db() as conn:
        rows = conn.execute("SELECT * FROM shifts ORDER BY id").fetchall()
    return jsonify({"items": to_items(rows)})


@app.post("/api/admin/shifts")
@api_guard("ADMIN")
def shifts_create():
    d = request.get_json(silent=True) or {}
    try:
        name = str(d.get("name", "")).strip(); st = str(d.get("start_time", "")).strip(); et = str(d.get("end_time", "")).strip(); g = int(d.get("grace_minutes", 15)); time.fromisoformat(st); time.fromisoformat(et)
    except Exception:
        return jsonify({"message": "Invalid shift payload"}), 400
    with conn_db() as conn:
        try:
            cur = conn.execute("INSERT INTO shifts (name,start_time,end_time,grace_minutes) VALUES (?,?,?,?)", (name, st, et, g)); conn.commit()
        except sqlite3.IntegrityError as e:
            return jsonify({"message": str(e)}), 400
    return jsonify({"message": "Shift created", "id": int(cur.lastrowid)})


@app.put("/api/admin/shifts/<int:shift_id>")
@api_guard("ADMIN")
def shifts_update(shift_id):
    d = request.get_json(silent=True) or {}
    fields, params = [], []
    if "name" in d: fields.append("name=?"); params.append(str(d.get("name", "")).strip())
    if "start_time" in d: time.fromisoformat(str(d.get("start_time"))); fields.append("start_time=?"); params.append(str(d.get("start_time")))
    if "end_time" in d: time.fromisoformat(str(d.get("end_time"))); fields.append("end_time=?"); params.append(str(d.get("end_time")))
    if "grace_minutes" in d: fields.append("grace_minutes=?"); params.append(int(d.get("grace_minutes")))
    if not fields: return jsonify({"message": "No fields to update"}), 400
    params.append(shift_id)
    with conn_db() as conn:
        try:
            conn.execute(f"UPDATE shifts SET {', '.join(fields)} WHERE id=?", tuple(params)); conn.commit()
        except sqlite3.IntegrityError as e:
            return jsonify({"message": str(e)}), 400
    return jsonify({"message": "Shift updated"})


@app.delete("/api/admin/shifts/<int:shift_id>")
@api_guard("ADMIN")
def shifts_delete(shift_id):
    with conn_db() as conn:
        row = conn.execute("SELECT id FROM shifts WHERE id=?", (shift_id,)).fetchone()
        if not row:
            return jsonify({"message": "Shift not found"}), 404
        fallback = conn.execute("SELECT id FROM shifts WHERE id<>? ORDER BY id LIMIT 1", (shift_id,)).fetchone()
        if not fallback:
            return jsonify({"message": "Cannot delete the last remaining shift"}), 400
        fallback_id = int(fallback["id"])
        conn.execute("UPDATE users SET shift_id=? WHERE shift_id=?", (fallback_id, shift_id))
        conn.execute("DELETE FROM shifts WHERE id=?", (shift_id,))
        conn.commit()
    return jsonify({"message": "Shift deleted", "fallback_shift_id": fallback_id})

def recalc_attendance(conn, attendance_id):
    a = conn.execute("SELECT * FROM attendance WHERE id=?", (attendance_id,)).fetchone()
    if not a or not a["login_time"] or not a["logout_time"]:
        return
    p = profile_for_attendance(conn, int(a["user_id"]), a)
    m = calc_metrics(a["login_time"], a["logout_time"], p)
    conn.execute("UPDATE attendance SET total_hours=?,overtime=?,late_mark=?,status=?,updated_at=? WHERE id=?", (m["total_hours"], m["overtime"], m["late_mark"], m["status"], now_iso(), attendance_id))


@app.get("/api/admin/export.xlsx")
@api_guard("ADMIN")
def export_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({"message": "openpyxl is required for export"}), 500

    try:
        dfrom, dto, status, uid, sid, cid = parse_filters()
    except Exception as e:
        return jsonify({"message": str(e)}), 400
    where, params = attendance_where(dfrom, dto, status, uid, sid, cid)

    with conn_db() as conn:
        rows = conn.execute(
            f"""
            SELECT u.name employee,a.attendance_date,a.login_time,a.logout_time,a.total_hours,a.overtime,
                   a.late_mark,a.break_taken,a.status,
                   COALESCE(s.name,'General Shift') shift_name,COALESCE(c.name,'General') category_name
            FROM attendance a JOIN users u ON u.id=a.user_id
            LEFT JOIN shifts s ON s.id=u.shift_id LEFT JOIN employee_categories c ON c.id=u.category_id
            WHERE {where} ORDER BY a.attendance_date DESC,a.id DESC
            """,
            tuple(params),
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Employee", "Date", "Login", "Logout", "Hours", "OT", "Late", "Break Taken", "Status", "Shift", "Category"])
    for r in rows:
        ws.append([r["employee"], r["attendance_date"], r["login_time"], r["logout_time"], r["total_hours"], r["overtime"], r["late_mark"], r["break_taken"], r["status"], r["shift_name"], r["category_name"]])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"attendance_{dfrom.isoformat()}_{dto.isoformat()}.xlsx")


@app.get("/api/employee/export.xlsx")
@api_guard("EMPLOYEE")
def export_my_attendance_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({"message": "openpyxl is required for export"}), 500

    u = request.current_user
    try:
        dfrom, dto = resolve_date_range(request.args.get("from"), request.args.get("to"))
    except Exception as e:
        return jsonify({"message": str(e)}), 400

    with conn_db() as conn:
        rows = conn.execute(
            """
            SELECT attendance_date,login_time,logout_time,total_hours,overtime,late_mark,break_taken,status
            FROM attendance WHERE user_id=? AND attendance_date BETWEEN ? AND ?
            ORDER BY attendance_date DESC,id DESC
            """,
            (u["id"], dfrom.isoformat(), dto.isoformat()),
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "My Attendance"
    ws.append(["Date", "Login", "Logout", "Hours", "OT", "Late", "Break Taken", "Status"])
    for r in rows:
        ws.append([r["attendance_date"], r["login_time"], r["logout_time"], r["total_hours"], r["overtime"], r["late_mark"], r["break_taken"], r["status"]])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"my_attendance_{dfrom.isoformat()}_{dto.isoformat()}.xlsx")

restore_db_from_github()
init_db()

if __name__ == "__main__":
    host = os.getenv("FLASK_RUN_HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "8000"))
    if IS_PRODUCTION:
        try:
            from waitress import serve
        except ImportError as exc:
            raise RuntimeError("waitress is required when APP_ENV=production. Install with: pip install -r requirements.txt") from exc
        threads = int(os.getenv("WAITRESS_THREADS", "8"))
        serve(app, host=host, port=port, threads=threads)
    else:
        debug = os.getenv("FLASK_DEBUG", "0") == "1"
        app.run(host=host, port=port, debug=debug)
